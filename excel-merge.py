from openpyxl import Workbook
from openpyxl import load_workbook
import os

if __name__=="__main__":
    #=== CONFIG
    targetPath = './끝/'
    saveFileName = 'MERGE.xlsx'

    #=== New Excel
    saveWb = Workbook()
    saveWs1 = saveWb.worksheets[0]
    header1 = ['제조사', '모델', '세부모델', '등급']
    saveWs1.column_dimensions['A'].width = 30
    saveWs1.column_dimensions['B'].width = 30
    saveWs1.column_dimensions['C'].width = 30
    saveWs1.column_dimensions['D'].width = 50
    saveWs1.append(header1)

    #=== Copy
    files = os.listdir(targetPath)
    try:
        files.remove(saveFileName) # 자기가있으면 자기는 지우는거
    except:
        pass
    print(files)
    for file in files:
        print(">>> {} 파일 시작 ".format(file))
        wb_sheet = load_workbook(targetPath+file).active
        idx = 0
        # min_row is set to 2, to ignore the first row which contains the headers
        for row in wb_sheet.iter_rows(min_row=2):
            row_data = []
            for cell in row:
                row_data.append(cell.value)
            idx += 1
            saveWs1.append(row_data)
        print("\t>>> {} 줄 옮기기 완료".format(idx))
    saveWb.save(targetPath+saveFileName)