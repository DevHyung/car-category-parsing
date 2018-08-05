#-*-encoding:utf8-*-
import requests
from bs4 import BeautifulSoup
from selenium import webdriver
import time
from openpyxl import Workbook
from openpyxl import load_workbook
import os

def bobae_make_excel(name,dataList):
    """
        :호출예시 make_excel([ [1,2,3], [4,5,6] ]) or make_excel(2dArray)
        :param dataList:  [ data1, data2, data3 ] 꼴의 1차원 list를 가지는 2차원 list
        :return: 없
    """
    #=== CONFIG
    #name = name.replace('/','&')
    #FILENAME = name+".xlsx"
    #에서 한파일로 뽑기위해 코드변경0805
    FILENAME = "보배드림.xlsx"

    #=== SAVE EXCEL
    if os.path.exists('./'+FILENAME):
        wb = load_workbook(filename=FILENAME)
        ws1 = wb[wb.sheetnames[0]]
        for data in dataList:
            ws1.append(data)
        wb.save(FILENAME)
    else:
        wb = Workbook()
        ws1 = wb.worksheets[0]
        header1 = ['제조사', '모델', '세부모델','등급']
        ws1.column_dimensions['A'].width = 30
        ws1.column_dimensions['B'].width = 30
        ws1.column_dimensions['C'].width = 30
        ws1.column_dimensions['D'].width = 50
        ws1.append(header1)
        # data save

        for data in dataList:
            ws1.append(data)
        # end
        wb.save(FILENAME)
    try:
        print(">>> {} {} 저장 완료 ".format(dataList[0][0],dataList[0][1]))
    except:
        print("에러")
        print(dataList)
def bobae():
    #=== var
    #INDEX = int(input("추출할 제조사 인덱스 입력 ( 처음 1 ) ::"))-1
    #=== convert BS4
    html = requests.get('http://www.bobaedream.co.kr/mycar/mycar_list.php?sel_m_gubun=ALL')
    html.encoding = 'utf-8'  # 한글 인코딩으로 변환
    bs4 = BeautifulSoup(html.text,'lxml')


    #=== parsing
    driver = webdriver.Chrome('./chromedriver')
    driver.maximize_window()
    driver.get('http://www.bobaedream.co.kr/mycar/mycar_list.php?sel_m_gubun=ALL')
    time.sleep(2)

    # depth 1
    dds = bs4.find('div',class_='area-maker').find_all('dd')
    for dd in dds:
        driver.execute_script(dd.button['onclick'])
        dept1_Title = dd.span.get_text().strip()
        # depth 2
        time.sleep(3)
        bs4 = BeautifulSoup(driver.page_source, 'lxml')
        modelDiv = bs4.find('div', class_='area-model')
        dds2 = modelDiv.find_all('dd')
        print("______"*10)
        for dd2 in dds2:
            dataList = []

            driver.execute_script(dd2.button['onclick'])
            dept2_Title = dd2.span.get_text().strip()
            # depth 3
            time.sleep(3)
            bs4 = BeautifulSoup(driver.page_source, 'lxml')
            detailDiv = bs4.find('div', class_='area-detail')
            dds3 = detailDiv.find_all('dd')
            for dd3 in dds3:
                if dd3['style'] == '':
                    dept3_Title = dd3.label.get_text().strip()
                    dept3_Id = dd3.input['id']
                    while True:
                        try:
                            elem = driver.find_element_by_xpath('//*[@id="{}"]'.format(dept3_Id)).click()
                            break
                        except:
                            time.sleep(0.5)
                    # depth 4
                    time.sleep(3)
                    bs4 = BeautifulSoup(driver.page_source, 'lxml')
                    detailDiv = bs4.find('div', class_='area-grade')
                    dds4 = detailDiv.find_all('dd')
                    loopidx = 1
                    while True:
                        if len(dds4) == 0: # 한개있을때
                            while True:
                                try:
                                    elem = driver.find_element_by_xpath('//*[@id="{}"]'.format(dept3_Id)).click()
                                    break
                                except:
                                    time.sleep(0.5)

                            time.sleep(3)
                            bs4 = BeautifulSoup(driver.page_source, 'lxml')
                            detailDiv = bs4.find('div', class_='area-grade')
                            dds4 = detailDiv.find_all('dd')
                            loopidx += 1
                            if loopidx == 3:
                                break
                        else:
                            break
                    if loopidx == 3:
                        dataList.append([dept1_Title, dept2_Title, dept3_Title, ''])
                    for dd4 in dds4:
                        dataList.append([dept1_Title,dept2_Title,dept3_Title,dd4.label.get_text().strip()])

                    driver.find_element_by_xpath('//*[@id="{}"]'.format(dept3_Id)).click()

            bobae_make_excel(dept1_Title,dataList)
    # ~()
    driver.quit()

if __name__=="__main__":
    bobae()

